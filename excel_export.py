"""
export/excel_export.py
Excel workbook generation: styled dashboard sheets + raw data tables.
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from utils.helpers import safe_str
from utils.constants import DAYS, PERIOD_KEYS, PERIOD_TIMES


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _safe_sheet_name(name: str) -> str:
    """Return a valid Excel sheet name (≤31 chars, no forbidden characters)."""
    if not name:
        name = "Sheet"
    name = str(name).strip()
    for ch in ("\\", "/", "*", "?", ":", "[", "]"):
        name = name.replace(ch, "-")
    name = name.replace("\n", " ").replace("\r", " ") or "Sheet"
    return name[:31]


def _count_lines(text: str) -> int:
    if not text:
        return 0
    return max(1, str(text).count("\n") + 1)


# ---------------------------------------------------------------------------
# Dashboard grid sheet
# ---------------------------------------------------------------------------

def _write_grid_sheet(ws, schedule_df: pd.DataFrame | None, title: str = "Timetable") -> None:
    """Write a styled timetable grid to worksheet *ws*."""
    days_order = DAYS
    periods_order = PERIOD_KEYS

    # Build grid: day → period → [(type, text), …]
    grid = {d: {p: [] for p in periods_order} for d in days_order}
    if schedule_df is not None and not schedule_df.empty:
        for _, r in schedule_df.iterrows():
            d = str(r.get("Day", "")).strip()
            p = str(r.get("Period", "")).strip()
            if d in grid and p in grid[d]:
                typ = str(r.get("Type", "")).strip().upper()
                cname = str(r.get("Course Name", "")).strip()
                instr = str(r.get("Instructor", "")).strip()
                room = str(r.get("Room", "")).strip()
                g = r.get("Group", "")
                gtxt = f"G{g}" if str(g).strip() else ""
                block = f"{typ}\n{cname}\n{instr}\n{room}"
                if gtxt:
                    block += f"\n{gtxt}"
                grid[d][p].append((typ, block))

    # Styles
    thin = Side(style="thin", color="8A8A8A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F2D3D")
    day_fill = PatternFill("solid", fgColor="243B53")
    lecture_fill = PatternFill("solid", fgColor="E8F1FF")
    lab_fill = PatternFill("solid", fgColor="E9FFF0")
    mixed_fill = PatternFill("solid", fgColor="FFF7E6")
    free_fill = PatternFill("solid", fgColor="F3F4F6")
    header_font = Font(color="FFFFFF", bold=True)
    day_font = Font(color="FFFFFF", bold=True)
    normal_font = Font(color="111827")
    type_font = Font(color="111827", bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_cell = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Title row
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(periods_order))
    ws["A1"].fill = header_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = align_center
    ws["A1"].border = border
    ws.row_dimensions[1].height = 30

    # Header row
    ws["A2"] = "Day / Period"
    ws["A2"].fill = header_fill
    ws["A2"].font = header_font
    ws["A2"].alignment = align_center
    ws["A2"].border = border

    for j, p in enumerate(periods_order, start=2):
        cell = ws.cell(row=2, column=j, value=f"{p}\n{PERIOD_TIMES[p]}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border
    ws.row_dimensions[2].height = 35

    # Data rows
    for row_i, day in enumerate(days_order, start=3):
        ws.cell(row=row_i, column=1, value=day).fill = day_fill
        ws.cell(row=row_i, column=1).font = day_font
        ws.cell(row=row_i, column=1).alignment = align_center
        ws.cell(row=row_i, column=1).border = border

        max_lines = 1
        for col_j, period in enumerate(periods_order, start=2):
            items = grid[day][period]
            if items:
                text = "\n---\n".join(block for (_, block) in items)
                types = {typ for (typ, _) in items}
                fill = (
                    lecture_fill if types == {"LECTURE"}
                    else lab_fill if types == {"LAB"}
                    else mixed_fill
                )
            else:
                text = "—"
                fill = free_fill

            cell = ws.cell(row=row_i, column=col_j, value=text)
            cell.fill = fill
            cell.font = normal_font
            cell.alignment = align_cell
            cell.border = border
            max_lines = max(max_lines, _count_lines(text))

        ws.row_dimensions[row_i].height = max(30, min(120, 15 * max_lines))

    _style_sheet_for_dashboard(ws, periods_order)


def _style_sheet_for_dashboard(ws, periods_order: list[str]) -> None:
    ws.freeze_panes = "B3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:2"
    ws.print_title_cols = "A:A"
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.column_dimensions["A"].width = 18
    for j in range(2, 2 + len(periods_order)):
        ws.column_dimensions[get_column_letter(j)].width = 36


def _write_raw_table(wb: Workbook, sheet_name: str, df: pd.DataFrame | None) -> None:
    """Write a styled raw table sheet to *wb*."""
    if df is None or df.empty:
        return
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="8A8A8A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=str(col))
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = align

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = border
            cell.alignment = align

    for j in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def df_to_excel_bytes(dfs: dict[str, pd.DataFrame]) -> bytes:
    """Simple multi-sheet export (raw tables, no extra styling)."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in dfs.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=name, index=False)
    return buf.getvalue()


def export_full_workbook_bytes(
    schedule_df: pd.DataFrame | None,
    summary_df: pd.DataFrame | None,
    conflicts_df: pd.DataFrame | None,
) -> bytes:
    """
    Build a rich workbook:
    - "All Courses" grid sheet
    - Per-doctor grid sheets
    - Per-assistant grid sheets
    - Raw data sheets (Schedule_Rows, Summary, Conflicts)
    """
    wb = Workbook()

    # --- All-courses sheet
    ws_all = wb.active
    ws_all.title = "All Courses"
    _write_grid_sheet(ws_all, schedule_df, title="University Timetable — All Courses")

    # --- Per-doctor sheets
    if schedule_df is not None and not schedule_df.empty and "Instructor" in schedule_df.columns:
        for instr_name, g in schedule_df.groupby("Instructor"):
            name = str(instr_name).strip() or "Unknown"
            sheet_title = _safe_sheet_name(f"Dr-{name}")
            if sheet_title in wb.sheetnames:
                sheet_title = _safe_sheet_name(f"Dr-{name}-2")
            ws = wb.create_sheet(sheet_title)
            _write_grid_sheet(ws, g, title=f"Doctor: {name}")

    # --- Raw tables
    _write_raw_table(wb, "Schedule_Rows", schedule_df)
    _write_raw_table(wb, "Summary", summary_df)
    _write_raw_table(wb, "Conflicts", conflicts_df)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
